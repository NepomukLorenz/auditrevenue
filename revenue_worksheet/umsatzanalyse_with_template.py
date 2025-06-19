import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font, Border, PatternFill, Alignment, Protection

from pathlib import Path
from shutil import copyfile
from typing import Union



__all__ = ["create_arbeitspapier_from_template_with_sections"]

def _write_block(ws, start_row: int, df: pd.DataFrame) -> None:
    """Schreibt eine 12-Zeilen-Tabelle (Umsatz, Materialaufwand) ab `start_row`."""
    for i in range(12):
        r = start_row + i
        ws.cell(r, 2, df.iloc[i]["Umsatz"])
        ws.cell(r, 3, df.iloc[i]["Materialaufwand"])

def _test_df(df: pd.DataFrame) -> bool:
    if len(df) != 12 or not {"Umsatz", "Materialaufwand"} <= set(df.columns):
            raise ValueError("df_current muss 12 Zeilen und die Spalten "
                             "'Umsatz', 'Materialaufwand' enthalten")
    return True

def _clear_worksheet_below(ws, start_row: int):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            cell.comment = None
            cell.font = Font()
            cell.border = Border()
            cell.fill = PatternFill(fill_type=None)
            cell.number_format = "General"
            cell.protection = Protection()
            cell.alignment = Alignment()
    ws.delete_rows(start_row, ws.max_row - start_row + 1)
    ws._charts = [
        ch for ch in ws._charts
        if getattr(ch.anchor, "_from", None) and ch.anchor._from.row < start_row - 1
    ]
    if hasattr(ws, "_images"):
        ws._images = [
            img for img in ws._images
            if getattr(img.anchor, "_from", None) and img.anchor._from.row < start_row - 1
        ]
    ws.merged_cells.ranges = [
        rng for rng in ws.merged_cells.ranges if rng.min_row < start_row
    ]
    cf = ws.conditional_formatting
    to_keep = []

    for entry in list(cf._cf_rules.values()):
        for rule in entry:
            try:
                min_col, min_row, max_col, max_row = range_boundaries(str(rule.sqref))
                if max_row < start_row:
                    to_keep.append((rule.sqref, rule))
            except Exception:
                continue

    cf._cf_rules.clear()

    for sqref, rule in to_keep:
        if isinstance(rule, Rule):
            cf.add(sqref, rule)



def create_arbeitspapier_from_template_with_sections(
        df_list: list[tuple[pd.DataFrame, pd.DataFrame]],
        template_path: Union[str, Path],
        output_path: Union[str, Path] = "arbeitspapier.xlsx",
        mus_sample: Union[pd.Series, pd.DataFrame] = None,
        cut_off_sample: Union[pd.Series, pd.DataFrame] = None,
    ) -> None:
    """create_arbeitspapier_from_template_with_sections"""

    template_path = Path(template_path)
    output_path = Path(output_path)
    if output_path.resolve() == template_path.resolve():
        output_path = output_path.with_stem(output_path.stem + "_out")

    copyfile(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb.worksheets[1]

    start_row = 20
    for i, (df_current, df_prior) in enumerate(df_list):
        _test_df(df_current)
        _test_df(df_prior)  

        _write_block(ws, start_row, df_current)
        start_row = start_row + 20
        _write_block(ws, start_row, df_prior)
        start_row = start_row + 24

    for ws_iter in wb.worksheets:
        for ch in ws_iter._charts:
            ch.style = 1
            ch.roundedCorners = False
            
    last_row = start_row -5
    _clear_worksheet_below(ws, last_row)

    _add_sample_on_new_sheet(ws=wb.worksheets[2], sample=mus_sample)

    _add_sample_on_new_sheet(ws=wb.worksheets[3], sample=cut_off_sample)

    wb.save(output_path)


def _add_sample_on_new_sheet(ws, sample: Union[pd.DataFrame, pd.Series]):
    """
    Fügt das gegebene Sample (Series oder DataFrame) ab Zelle A18 in das zweite Worksheet ein.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Ein beliebiges Tabellenblatt aus der Arbeitsmappe.
    sample : pd.Series oder pd.DataFrame
        Die zu schreibende Stichprobe.
    """
    target_ws = ws

    if isinstance(sample, pd.Series):
        df = sample.to_frame(name=sample.name or "value")
    elif isinstance(sample, pd.DataFrame):
        df = sample.copy()
    elif sample is None:
        pass
    else:
        raise TypeError("Sample muss eine pandas Series oder DataFrame sein.")

    start_row = 18 # ab Zeile 18, Spalte A
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=df is sample, header=True), start=start_row
    ):
        for c_idx, value in enumerate(row, start=1):
            target_ws.cell(row=r_idx, column=c_idx, value=value)

    from openpyxl.styles import Font
    hdr_font = Font(bold=True)
    for c_idx in range(1, df.shape[1] + (0 if isinstance(sample, pd.Series) else 1) + 1):
        target_ws.cell(row=start_row, column=c_idx).font = hdr_font